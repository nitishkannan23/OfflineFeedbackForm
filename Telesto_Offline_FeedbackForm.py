# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
from tkinter import messagebox
from tkinter.scrolledtext import ScrolledText
import xlsxwriter

# globally declare wb and sheet variable 
  
file = 'Feedback_Form_Telesto.xlsx'

try:
    workbook = xlsxwriter.Workbook('./'+file)
    worksheet = workbook.add_worksheet()
    workbook.close()
except:
    pass

# opening the existing excel file 
wb = load_workbook('./'+file) 
  
# create the sheet object 
sheet = wb.active 
  
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 100
    sheet.column_dimensions['H'].width = 100

  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Company"
    sheet.cell(row=1, column=3).value = "Designation"
    sheet.cell(row=1, column=4).value = "Email ID"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Quality of Service"
    sheet.cell(row=1, column=7).value = "Fields to Focus"
    sheet.cell(row=1, column=8).value = "Remarks"
  
  
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the course_field box 
    company_field.focus_set() 
   
# Function to set focus 
def focus2(event): 
    # set focus on the sem_field box 
    designation_field.focus_set() 

def focus3(event): 
    # set focus on the email_id_field box 
    email_id_field.focus_set() 
    
def focus4(event): 
    # set focus on the contact_no_field box 
    contact_no_field.focus_set() 
    
# Function to set focus 
def focus6(event): 
    # set focus on the address_field box 
    popupmenu1_field.focus_set() 
    
# Function to set focus 
def focus7(event): 
    # set focus on the address_field box 
    popupmenu2_field.focus_set() 
    
    # Function to set focus 
def focus8(event): 
    # set focus on the address_field box 
    text1_field.focus_set() 
  
  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    name_field.delete(0, END) 
    company_field.delete(0, END) 
    designation_field.delete(0, END) 
    email_id_field.delete(0, END) 
    contact_no_field.delete(0, END) 
    popupmenu1_field.delete(0, END) 
    popupmenu2_field.delete(0, END) 
    text1.delete(1.0, END) 
  
  
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" or
        company_field.get() == "" or
        designation_field.get() == "" or
        email_id_field.get() == "" or
        contact_no_field.get() == "" or
        var1.get() == "" or
        var2.get() == "" or
        text1.get(1.0, END) == ""): 
              
        print("empty input")
        messagebox.showerror("Error", "Please Fill all Fields!")
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = company_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = designation_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = email_id_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = var1.get() 
        sheet.cell(row=current_row + 1, column=7).value = var2.get() 
        sheet.cell(row=current_row + 1, column=8).value = text1.get("1.0", "end-1c") 
  
        # save the file 
        wb.save('./'+file) 
  
        # set focus on the name_field box 
        name_field.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
    
    # set the background colour of GUI window 
    root.configure(background='#ceccb8', width=50,height=50) 
  
    # set the title of GUI window 
    root.title("Telesto Feedback Form") 
  
    # set the configuration of GUI window 
    root.geometry("1920x1080") 
  
    excel() 
    
    string1 = '\n'
    

    # create a Form label 
    heading = Label(root, text="Telesto Energy - Feedback Form", bg="#ceccb8", fg="#000066", font='verdana 22 bold').grid(row=0, column=1) 
    break1= Label(root, text=string1, bg="#ceccb8").grid(row=1, column=1) 
                    
    # create a Name label 
    name = Label(root, text="Name", bg="#ceccb8", fg="#000066", font='verdana 12 bold').grid(row=2, column=0)              
    break2= Label(root, text=string1, bg="#ceccb8").grid(row=3, column=0) 
  
    # create a Course label 
    company = Label(root, text="Company", bg="#ceccb8", fg="#000066", font='verdana 12 bold').grid(row=4, column=0)  
    break3= Label(root, text=string1, bg="#ceccb8").grid(row=5, column=0)  
  
    # create a Semester label 
    designation = Label(root, text="Designation", bg="#ceccb8", fg="#000066", font='verdana 12 bold').grid(row = 6, column = 0)
    break4= Label(root, text=string1, bg="#ceccb8").grid(row=7, column=0)  
    # create a Form No. lable 
    email = Label(root, text="Email ID.", bg="#ceccb8", fg="#000066", font='verdana 12 bold').grid(row=8, column=0)  
    break5= Label(root, text=string1, bg="#ceccb8").grid(row=9, column=0)  
                  
    # create a Contact No. label 
    contact_no = Label(root, text="Contact No.", bg="#ceccb8", fg="#000066", font='verdana 12 bold').grid(row=10, column=0)  
    break6= Label(root, text=string1, bg="#ceccb8").grid(row=11, column=0) 
                  
    # create a address label     
    var1= StringVar(root)
    var1.set('Average') # set the default option
    choices1 = { 'Bad','Average','Good','Excellent'}
    popupmenu1 = OptionMenu(root, var1, *choices1)
    Label(root, text="Quality of Service", bg="#ceccb8", fg="#000066", font='verdana 12 bold').grid(row=12, column=0) 
    break8= Label(root, text=string1, bg="#ceccb8").grid(row=13, column=0)  
                  
    var2= StringVar(root)
    var2.set('Oil and Gas Consulting') # set the default option
    choices2 = { 'Oil and Gas Consulting','AI and Big Data Analytics','Portfolio Management','Research and Development'}
    popupmenu2 = OptionMenu(root, var2, *choices2)
    Label(root, text="Fields to Focus", bg="#ceccb8", fg="#000066", font='verdana 12 bold').grid(row=14, column=0) 
    break9= Label(root, text=string1, bg="#ceccb8").grid(row=15, column=0) 
    
    text1 = ScrolledText(root, width  = 60, height = 10)
    remarks = Label(root, text="Remarks", bg="#ceccb8", fg="#000066", font='verdana 12 bold').grid(row=16, column=0) 
    break10= Label(root, text=string1, bg="#ceccb8").grid(row=17, column=0)  
  
    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    company_field = Entry(root) 
    designation_field = Entry(root) 
    email_id_field = Entry(root) 
    contact_no_field = Entry(root) 
    popupmenu1_field = Entry(root) 
    popupmenu2_field = Entry(root)
    text1_field = Entry(root)
    
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    name_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    company_field.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    designation_field.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed 
    # then call the focus4 function 
    email_id_field.bind("<Return>", focus4) 
  
    # whenever the enter key is pressed 
    # then call the focus5 function 
    contact_no_field.bind("<Return>", focus6) 
  
  
    popupmenu1_field.bind("<Return>", focus7) 
    
    popupmenu2_field.bind("<Return>", focus8) 
    
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    name_field.grid(row=2, column=1, ipadx="200",ipady="10") 
    company_field.grid(row=4, column=1, ipadx="200", ipady="10") 
    designation_field.grid(row=6, column=1, ipadx="200", ipady="10") 
    email_id_field.grid(row=8, column=1, ipadx="200", ipady="10") 
    contact_no_field.grid(row=10, column=1, ipadx="200", ipady="10") 
    popupmenu1.grid(row=12, column=1, ipadx="100", ipady="8") 
    popupmenu2.grid(row=14, column=1, ipadx="75", ipady="8") 
    text1.grid(row=16, column=1) 
  
    # call excel function 
    excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="white", 
                            bg="#000066", command=insert) 
    submit.grid(row=28, column=1, ipadx="25", ipady="10") 
  
    # start the GUI 
    root.mainloop() 